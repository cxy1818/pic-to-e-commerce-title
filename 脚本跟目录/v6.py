import os
import base64
import json
import requests
import re
from openpyxl import Workbook
with open("./模型调用参数/固定前缀.txt", "r", encoding="utf-8") as f:
    TITLE_PREFIX = f.read().strip()
with open("./模型调用参数/固定后缀.txt", "r", encoding="utf-8") as f:
    TITLE_SUFFIX = f.read().strip()
with open("./模型调用参数/KEY.txt", "r", encoding="utf-8") as f:
    API_KEY = f.read().strip()
with open("./模型调用参数/URL.txt", "r", encoding="utf-8") as f:    
    API_URL = f.read().strip()
with open("./模型调用参数/MODEL.txt", "r", encoding="utf-8") as f:
    MODEL = f.read().strip()
FOLDER = r"./待识别的图"
with open("./模型调用参数/提示词.txt", "r", encoding="utf-8") as f:
    PROMPT = f.read().strip()
with open("./模型调用参数/生成几条.txt", "r", encoding="utf-8") as f:
    GENERATE_COUNT = int(f.read().strip())

exts = [".jpg", ".jpeg", ".png", ".webp", ".bmp"]

headers = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {API_KEY}"
}
wb = Workbook()
ws = wb.active
ws.title = "图片标题报告"
current_row = 1  
processed_prefixes = set()
file_groups = {}

for filename in os.listdir(FOLDER):
    if not any(filename.lower().endswith(e) for e in exts):
        continue
    print(f"检查文件: {filename}")  
    match = re.match(r'^(.*?)[-_](\d+)\.[A-Za-z0-9]+$', filename)
    if match:
        prefix = match.group(1)
        number = int(match.group(2))
        print(f"匹配成功: 前缀='{prefix}', 数字={number}")  
        if prefix not in file_groups:
            file_groups[prefix] = []
        file_groups[prefix].append((number, filename))
    else:
        print(f"未匹配分组规则，单独处理: {filename}")
        if filename not in file_groups:
            file_groups[filename] = [(0, filename)]
for prefix, files in file_groups.items():
    files.sort(key=lambda x: x[0])
    first_number, first_filename = files[0]
    print(f"处理 {prefix} 系列的第一张图片: {first_filename}")
    path = os.path.join(FOLDER, first_filename)
    with open(path, "rb") as f:
        img_b64 = base64.b64encode(f.read()).decode()
    full_prompt = f"{PROMPT}\n生成 {GENERATE_COUNT} 条标题分行输出"
    data = {
        "model": MODEL,
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{img_b64}"
                        }
                    },
                    {
                        "type": "text",
                        "text": full_prompt
                    }
                ]
            }
        ]
    }
    response = requests.post(API_URL, headers=headers, data=json.dumps(data))
    print("HTTP 状态码：", response.status_code)
    print("返回内容：", response.text)
    result = response.json()
    try:
        answer = result["choices"][0]["message"]["content"]
    except:
        answer = json.dumps(result, ensure_ascii=False)   
    lines = answer.split("\n")
    lines = [l.strip() for l in lines if l.strip()]  
    lines = [re.sub(r'^\d+[\.、]?\s*', '', l) for l in lines]     
    lines = [f"{TITLE_PREFIX}{l}{TITLE_SUFFIX}" for l in lines]
    for title in lines:
        filename_without_ext = os.path.splitext(first_filename)[0]
        ws.cell(row=current_row, column=1, value=filename_without_ext)  
        ws.cell(row=current_row, column=2, value=title)       
        current_row += 1
    if len(files) > 1:
        print(f"{prefix} 系列识别完成，跳过后续图片: {[f[1] for f in files[1:]]}")
    print()
from datetime import datetime
output_path = f"以图片生标题_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb.save(output_path)
print("表格已生成在脚本同路径文件夹")